import csv

'''*****************************************
STEP: 1
Reading all row into memory to process later
CSV File from: https://simplemaps.com/data/us-cities
*****************************************'''
#create list to store row object
city_list =[]
#will change to False and skip first row with title informaiton
skip_first_row = True

#Read CSV file into memory
with open('uscities.csv') as csvfile:
    readCSV = csv.reader(csvfile, delimiter=',')
    for row in readCSV:
        #skip first row
        if skip_first_row:
            skip_first_row = False
            continue
        #store each row in a dictionary
        cityInfo ={
            'city' : row[0],
            'state_id': row[2],
            'state_name': row[3],
            'city_population':row[10],
            'military_base': row[13],
            'time_zone' : row[15]
        }

        #print(cityInfo)
        city_list.append(cityInfo)

#print(len(city_list))
#print(city_list[0]['city'])

#get list of all city id
all_state_id = []
for id in city_list:
    #check in city id is in list. If not then add to list
    if id['state_id'] not in all_state_id:
        all_state_id.append(id['state_id'])
#print(len(all_state_id))



'''**********************************************************
STEP: 2
check if current directory has a folder called "stateFiles"
If not then create one.
**********************************************************'''
import os
excel_folder_name = os.getcwd()+'\stateFiles'

if os.path.exists(excel_folder_name) == False or os.path.isdir(excel_folder_name) == False:

    try:
        os.mkdir(excel_folder_name)
    except OSError:
        print("Creation of the directory %s failed" % excel_folder_name)
    else:
        print("Successfully created the directory %s " % excel_folder_name)



'''**********************************************************
STEP: 3
    1. Use Openpyxl to create workbook for each state_id
    2. Create table in each workbook containing all information for that state
    3. Highlight rows for cities with Military Base
**********************************************************'''


from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Color, PatternFill, Font, Border
import datetime

#FUNCITON WILL BE USED TO CREATE NEW FILES ---------------------------------------
def newFile(id):
    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')
    #create new workbook with Openpyxl
    wb = Workbook()
    #using firt worksheet open in excel "Sheet1"
    ws = wb.active
    #tracking to see how many rows are inserted
    rowcount = 1
    #example number 1 to write value to cell
    ws['A1'] = 'State Name'
    #example number 2 to write value to cell
    ws.cell(row=1,column=2,value='city')
    ws.cell(row=1, column=3, value='time_zone')
    ws.cell(row=1, column=4, value='City Population')
    ws.cell(row=1, column=5, value='Military Base')
    #change title name of workbook
    ws.title ='City Data'

    #reset flag to Non Military state
    mil = 'NON_Military_State'

    for row in city_list:
        if row['state_id'] == id:
            rowcount +=1

            #print(row['city'])
            #print(row['state_name'])
            #print(row['city_population'])
            #print(row['military_base'])
            #print(row['time_zone'])

            ws.append([row['state_name'], row['city'], row['time_zone'], row['city_population'], row['military_base']])
            if row['military_base'] == 'TRUE':
                #change flag if state has military base
                mil = "Military_State"
                for x in range (1,len(row)):
                    ws.cell(row=rowcount, column=x).fill = redFill


    #print(rowcount)

    tab = Table(displayName="Table1", ref="A1:E"+str(rowcount))

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    wb.create_sheet('Cover_Page',index=0)

    #--------- Create new page in workbook -----------------------

    coverPage = wb['Cover_Page']

    coverPage['A1'] = 'Created BY:'
    coverPage['A2'] = "Today's date:"
    coverPage['A3'] = "Military State"
    coverPage['B1'] = 'Juan Nadal'
    coverPage['B2'] = datetime.datetime.now()
    if mil == 'Military_State':
        coverPage['B3'] = 'YES'
    else:
        coverPage['B3'] = 'NO'


    #save file and name the state_id and if Military state
    wb.save(excel_folder_name+'\\'+id+'_'+mil+'.xlsx')



#END OF FUNCTION ---------------------------------------------------------------

for i in all_state_id:
    newFile(i)
    #ONLY CREATE ONE FILE WHEN TESTING
    break



__name__ = "__main__"

