from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Color, PatternFill, Font, Border
import datetime

#FUNCITON WILL BE USED TO CREATE NEW FILES ---------------------------------------
def newFile(id,city_list,excel_folder_name):
    redFill = PatternFill(start_color='FFFF0000',
                          end_color='FFFF0000',
                          fill_type='solid')
    #create new workbook with Openpyxl
    wb = Workbook()
    #using firt worksheet open in excel "Sheet1"
    ws = wb.active
    #tracking to see how many rows are inserted
    rowcount = 1
    #example number 1 to write value to cell
    ws['A1'] = 'State Name'
    #example number 2 to write value to cell
    ws.cell(row=1,column=2,value='city')
    ws.cell(row=1, column=3, value='time_zone')
    ws.cell(row=1, column=4, value='City Population')
    ws.cell(row=1, column=5, value='Military Base')
    #change title name of workbook
    ws.title ='City Data'

    #reset flag to Non Military state
    mil = 'NON_Military_State'

    for row in city_list:
        if row['state_id'] == id:
            rowcount +=1

            #print(row['city'])
            #print(row['state_name'])
            #print(row['city_population'])
            #print(row['military_base'])
            #print(row['time_zone'])

            ws.append([row['state_name'], row['city'], row['time_zone'], row['city_population'], row['military_base']])
            if row['military_base'] == 'TRUE':
                #change flag if state has military base
                mil = "Military_State"
                for x in range (1,len(row)):
                    ws.cell(row=rowcount, column=x).fill = redFill


    #print(rowcount)

    tab = Table(displayName="Table1", ref="A1:E"+str(rowcount))

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    wb.create_sheet('Cover_Page',index=0)

    #--------- Create new page in workbook -----------------------

    coverPage = wb['Cover_Page']

    coverPage['A1'] = 'Created BY:'
    coverPage['A2'] = "Today's date:"
    coverPage['A3'] = "Military State"
    coverPage['B1'] = 'Juan Nadal'
    coverPage['B2'] = datetime.datetime.now()
    if mil == 'Military_State':
        coverPage['B3'] = 'YES'
    else:
        coverPage['B3'] = 'NO'


    #save file and name the state_id and if Military state
    wb.save(excel_folder_name+'\\'+id+'_'+mil+'.xlsx')



#END OF FUNCTION ---------------------------------------------------------------